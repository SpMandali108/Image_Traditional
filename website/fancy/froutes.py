import json
from typing import Counter
from flask import Blueprint, render_template, request, redirect, url_for, session, jsonify,flash
from datetime import datetime, timedelta
from bson.objectid import ObjectId
import os

from .fservices import *
from .fmodels import *
from ..general.db import *

from website.fancy.fcycle import fancy_cycles

from website.fancy.fcycle import (
    get_active_cycle,
    get_selected_cycle_id,
    get_all_cycles,
    get_cycle_by_id,
    get_selected_cycle,
    is_selected_cycle_locked,
    set_selected_cycle,
    create_cycle,
    end_cycle,
    reactivate_cycle,
    get_active_collection,
    get_selected_collection
)
from .fnormalizer import normalize_cycle_bookings, clean_school_name
from .fanalytics import (
    compute_dashboard_metrics,
    get_category_drilldown,
    get_product_drilldown,
    get_school_drilldown,
    get_customer_drilldown,
    get_month_drilldown,
    get_day_drilldown,
    get_date_drilldown,
    get_cycle_drilldown,
    safe_parse_date
)
from .fai import (
    generate_dashboard_ai_insights,
    explain_anomaly_spikes,
    generate_festival_forecast_insights,
    generate_drilldown_ai_insights
)

fancy = Blueprint('fancy', __name__)

# ------------------ MAIN ------------------


@fancy.route('/fancy', methods=['GET', 'POST'])
def fbook():
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))

    if request.method == 'POST':
        data = request.get_json(silent=True)

        if is_selected_cycle_locked():
            return jsonify({
        "error": "Cycle is locked"
    }), 403

        if not data:
            return jsonify({"error": "Invalid JSON"}), 400

        # Normalize keys
        data = {k.lower(): v for k, v in data.items()}

        mobile = str(data.get('mobile', '')).strip()
        if not mobile or len(mobile) != 10:
            return jsonify({"error": "Invalid mobile"}), 400

        collection = get_selected_collection()
        school = data.get('school', '').strip().title()
        costume = data.get('costume', '').strip().title()

        booking_data = {
    'name': data.get('name', '').strip().title(),
    'mobile': mobile,
    'address': data.get('address', '').strip().title(),
    'school': data.get('school', '').strip().title(),
    'start_date': data.get('start_date', ''),
    'end_date': data.get('end_date', ''),
    'price': float(data.get('price', 0)),
    'costume': data.get('costume', '').strip().title(),
    'details': data.get('details', '').strip(),
    'timestamp': datetime.utcnow()
}

        customer_data = {
            'name': booking_data['name'],
            'mobile': mobile,
            'address': booking_data['address'],
            'school': booking_data['school'],
            'updated_at': datetime.utcnow()
        }

        fcustomers.update_one(
            {'mobile': mobile},
            {
                '$set': customer_data,
                '$setOnInsert': {'created_at': datetime.utcnow()}
            },
            upsert=True
        )

        # Add school to School_Master if new
        if school:
            db.School_Master.update_one(
                {"name": school},
                {"$setOnInsert": {"name": school}},
                upsert=True
            )

        # Add costume category to Costume_Category_Master if new
        if costume:
            db.Costume_Category_Master.update_one(
                {"name": costume},
                {"$setOnInsert": {"name": costume}},
                upsert=True
            )

        collection.insert_one(booking_data)

        return jsonify({'status': 'success'}), 200

    # ✅ GET request — data is NOT used here
    schools = sorted(
    [x["name"] for x in db.School_Master.find({}, {"name": 1})]
)

    costumes = sorted(
        [x["name"] for x in db.Costume_Category_Master.find({}, {"name": 1})]
    )

    return render_template(
        "fancy/fancy.html",
        schools=schools,
        costumes=costumes
    )

@fancy.route("/fancy_listing",methods=['GET','POST'])
def flisting():
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))
    
    collection = get_selected_collection()
    fbookings = list(collection.find())

    return render_template("fancy/fancy_listing.html",fbookings = fbookings)





@fancy.route('/delete_booking/<id>', methods=['POST'])
def delete_booking(id):

    if not session.get('logged_in'):
        return jsonify(success=False)

    collection = get_selected_collection()

    collection.delete_one({
        '_id': ObjectId(id)
    })

    return jsonify(success=True)

@fancy.route('/update_booking', methods=['POST'])
def update_booking():

    try:
        data = request.json

        current_app.logger.info(f"DATA = {data}")

        collection = get_selected_collection()

        collection.update_one(
    {'_id': ObjectId(data['id'])},
    {
        '$set': {
            'name': data['name'],
            'mobile': data['mobile'],
            'address': data['address'],
            'school': data['school'],
            'costume': data['costume'],
            'details': data['details'],
            'price': int(float(data['price'])),
            'start_date': data['start_date'],
            'end_date': data['end_date']
        }
    }
)

        return jsonify(success=True)

    except Exception as e:
        current_app.logger.error(f"ERROR: {e}")
        return jsonify(success=False, message=str(e))

@fancy.route("/get_customer")
def get_customer():
    mobile = request.args.get("mobile")

    customer = fcustomers.find_one(
        {"mobile": mobile},     # lowercase mobile
        {"_id": 0}
    )

    if customer:
        return jsonify({"exists": True, "data": customer})

    return jsonify({"exists": False})

@fancy.route('/fancy_calendar', methods=['GET', 'POST'])
def fancy_calendar():
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))

    today = datetime.now().date()
    selected_date = request.args.get('date')

    # ---------- HANDLE TAKEN / RETURNED ----------
    if request.method == 'POST':

        actions_raw = request.form.get('actions')

        if not actions_raw:
            return jsonify(success=False)

        actions = json.loads(actions_raw)

        for act in actions:

            bid = act['bookingId']
            field = act['field']
            cycle_id = act.get('cycleId')

            if not cycle_id:
                continue

            cycle = get_cycle_by_id(cycle_id)

            if not cycle:
                continue

            collection = db[
                cycle['collection_name']
            ]

            collection.update_one(
                {'_id': ObjectId(bid)},
                {'$set': {field: True}}
            )

        return jsonify(success=True)

    # ---------- FETCH ALL BOOKINGS ----------
    all_bookings = []

    for cycle in get_all_cycles():

        collection = db[
            cycle['collection_name']
        ]

        for b in collection.find():

            b['season'] = cycle['name']
            b['cycle_id'] = str(cycle['_id'])

            # Normalize dates
            for k in ['start_date', 'end_date']:

                v = b.get(k)

                if isinstance(v, datetime):

                    b[k] = v.strftime('%d-%m-%Y')

                elif isinstance(v, str):

                    formats = [
                        '%Y-%m-%d',
                        '%d-%m-%Y',
                        '%d-%m-%y'
                    ]

                    for fmt in formats:
                        try:
                            b[k] = datetime.strptime(
                                v,
                                fmt
                            ).strftime('%d-%m-%Y')
                            break
                        except:
                            pass

            all_bookings.append(b)

    # ---------- CALENDAR HIGHLIGHT DATES ----------
    booked_dates = set()

    for b in all_bookings:

        try:

            sd = datetime.strptime(
                b['start_date'],
                '%d-%m-%Y'
            ).date()

            ed = datetime.strptime(
                b['end_date'],
                '%d-%m-%Y'
            ).date()

            cur = sd

            while cur <= ed:

                booked_dates.add(
                    cur.strftime('%Y-%m-%d')
                )

                cur += timedelta(days=1)

        except:
            pass

    # ---------- BOOKINGS FOR SELECTED DATE ----------
    day_bookings = []

    if selected_date:

        sel = datetime.strptime(
            selected_date,
            '%Y-%m-%d'
        ).date()

        for b in all_bookings:

            try:

                sd = datetime.strptime(
                    b['start_date'],
                    '%d-%m-%Y'
                ).date()

                ed = datetime.strptime(
                    b['end_date'],
                    '%d-%m-%Y'
                ).date()

                if sd <= sel <= ed:
                    day_bookings.append(b)

            except:
                pass

    # ---------- UPCOMING & NOT RETURNED ----------
    upcoming = []
    not_returned = []

    for b in all_bookings:

        try:

            ed = datetime.strptime(
                b['end_date'],
                '%d-%m-%Y'
            ).date()

            if ed >= today:

                upcoming.append(b)

            elif ed < today and not b.get('returned'):

                not_returned.append(b)

        except:
            pass

    return render_template(
        'fancy/fancy_calendar.html',
        booked_dates=list(booked_dates),
        day_bookings=day_bookings,
        upcoming=upcoming,
        not_returned=not_returned,
        selected_date=selected_date,
        today=today.strftime('%Y-%m-%d')
    )

@fancy.route('/fancy_dashboard')
def fancy_dashboard():

    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))

    # -----------------------------
    # 1. CYCLE SELECTOR HANDLE
    # -----------------------------
    cycle_id = request.args.get('cycle_id')
    if cycle_id:
        set_selected_cycle(cycle_id)

    # -----------------------------
    # 2. SELECTED CYCLE DATA & BOOKINGS
    # -----------------------------
    collection = get_selected_collection()
    current_bookings = list(collection.find())

    # All bookings across all cycles (for longitudinal festival back-testing)
    all_bookings = []
    for cycle in get_all_cycles():
        c_col = db[cycle["collection_name"]]
        c_bookings = list(c_col.find())
        for b in c_bookings:
            b["season"] = cycle["name"]
        all_bookings.extend(c_bookings)

    # -----------------------------
    # 3. DETERMINISTIC ANALYTICS (Current Cycle)
    # -----------------------------
    metrics = compute_dashboard_metrics(current_bookings, all_bookings)
    primary_kpis = metrics["primary_kpis"]
    best_performers = metrics["best_performers"]
    charts = metrics["charts"]
    spikes = metrics["spikes"]
    top_schools_table = metrics["top_schools_table"]
    top_customers_table = metrics["top_customers_table"]

    # -----------------------------
    # 4. INVENTORY STOCK MAPPING
    # -----------------------------
    inventory_products = list(finventory.find())
    category_stock = {}
    total_stock = 0
    for p in inventory_products:
        cat = p.get("category", "General").strip().title()
        sizes = p.get("sizes", {})
        qty = 0
        if isinstance(sizes, dict):
            qty = sum(int(q) for q in sizes.values() if str(q).isdigit() or isinstance(q, (int, float)))
        category_stock[cat] = category_stock.get(cat, 0) + qty
        total_stock += qty

    # -----------------------------
    # 5. INDIAN FESTIVAL FORECAST CALENDAR
    # -----------------------------
    indian_events = [
        {"id": "republic_day", "name": "Republic Day", "month": 1, "day": 26, "category": "Freedom Fighters / Regional", "description": "National parade & school acts. High demand for Gandhi, Nehru, Bhagat Singh, Subhas Chandra Bose, and army/police uniforms."},
        {"id": "independence_day", "name": "Independence Day", "month": 8, "day": 15, "category": "Freedom Fighters / National Heroes", "description": "Independence Day assemblies. Highest demand for historic freedom fighter attire (Rani Laxmibai, Bhagat Singh, Gandhi, Nehru)."},
        {"id": "gandhi_jayanti", "name": "Gandhi Jayanti", "month": 10, "day": 2, "category": "Freedom Fighters / Khadi Attire", "description": "Birth anniversary of Mahatma Gandhi. High demand for dhotis, bald wigs, spectacles, and Nehru caps."},
        {"id": "teachers_day", "name": "Teachers' Day", "month": 9, "day": 5, "category": "Professional / Formal Costumes", "description": "Teachers' Day plays and roleplays. Demand for formal blazers, sarees, doctor, lawyer, and corporate uniforms."},
        {"id": "childrens_day", "name": "Children's Day", "month": 11, "day": 14, "category": "Cartoon Characters / Animals / Nehru", "description": "Children's Day events. High demand for cartoon characters (Doraemon, Mickey Mouse), animal onesies, and Chacha Nehru jackets."},
        {"id": "christmas", "name": "Christmas Concerts", "month": 12, "day": 25, "category": "Christmas / Angels / Santa Claus", "description": "School Christmas concerts. High demand for Santa Claus outfits, Elf costumes, Angel wings, and Shepherd robes."},
        {"id": "janmashtami", "name": "Krishna Janmashtami", "month": 9, "day": 3, "category": "Mythological (Krishna/Radha)", "description": "Krishna tableaus and dahi handi. Peak demand for Bal Krishna crown, flute, peacock feather, and Radha lehengas."},
        {"id": "navaratri", "name": "Navaratri Festival", "month": 10, "day": 12, "category": "Chaniya Choli / Kediyu", "description": "9 days of Garba. Huge demand for heavily-embroidered Chaniya Cholis, Kediyus, turbans, and oxidized ornaments."}
    ]

    import datetime as dt
    today_dt = datetime.now()
    forecast_calendar = []

    for ev in indian_events:
        year = today_dt.year
        ev_date = dt.datetime(year, ev["month"], ev["day"])
        if ev_date.date() < today_dt.date():
            ev_date = dt.datetime(year + 1, ev["month"], ev["day"])
        countdown = (ev_date.date() - today_dt.date()).days

        matching_bookings = []
        event_categories = Counter()
        event_costumes = Counter()
        for b in all_bookings:
            sd = safe_parse_date(b.get("start_date"))
            if sd:
                try:
                    ev_date_by_year = dt.date(sd.year, ev["month"], ev["day"])
                except ValueError:
                    ev_date_by_year = dt.date(sd.year, ev["month"], ev["day"] - 1)

                diff = (sd - ev_date_by_year).days
                if -3 <= diff <= 3:
                    costume = b.get("details", b.get("costume", "Unknown")).strip().title()
                    cat = b.get("costume", "General").strip().title()
                    event_categories[cat] += 1
                    event_costumes[costume] += 1
                    matching_bookings.append({
                        "name": b.get("name", "Unknown"),
                        "mobile": b.get("mobile", ""),
                        "costume": costume,
                        "date": sd.strftime("%d-%m-%Y"),
                        "price": b.get("price", 0),
                        "season": b.get("season", "Historical")
                    })

        top_cats = [c for c, _ in event_categories.most_common(2)]
        top_dresses = [f"{d} ({count})" for d, count in event_costumes.most_common(5)]
        dynamic_category = ", ".join(top_cats) if top_cats else ev["category"]
        dynamic_desc = f"Top historical costumes: {', '.join(top_dresses)}" if top_dresses else ev["description"]

        total_available_stock = sum(category_stock.get(cat, 0) for cat in top_cats) or category_stock.get("Bhagwan", 20)
        deficit = max(0, len(matching_bookings) - total_available_stock)

        forecast_calendar.append({
            "id": ev["id"],
            "name": ev["name"],
            "date": ev_date.strftime("%d-%m-%Y"),
            "countdown": countdown,
            "category": dynamic_category,
            "description": dynamic_desc,
            "spike_count": len(matching_bookings),
            "bookings": matching_bookings,
            "total_stock": total_available_stock,
            "deficit": deficit,
            "status": "Critical Deficit" if deficit > 10 else "Moderate Shortage" if deficit > 0 else "Stock Adequate"
        })

    forecast_calendar.sort(key=lambda x: x["countdown"])

    # -----------------------------
    # 6. AI ENRICHMENT (Cached & Privacy-Sanitized)
    # -----------------------------
    festival_ai_advice = generate_festival_forecast_insights(forecast_calendar)
    for ev in forecast_calendar:
        if ev["id"] in festival_ai_advice:
            ev["ai_advice"] = festival_ai_advice[ev["id"]].get("actionable_advice", "")
            ev["ai_outlook"] = festival_ai_advice[ev["id"]].get("demand_outlook", "Moderate")
            ev["ai_characters"] = festival_ai_advice[ev["id"]].get("key_characters", [])
        else:
            ev["ai_advice"] = "Monitor historical booking velocity and maintain buffer stock."
            ev["ai_outlook"] = "Steady"
            ev["ai_characters"] = []

    spike_explanations = explain_anomaly_spikes(spikes)
    spike_map = {sp["date"]: sp for sp in spike_explanations}
    for s in spikes:
        if s["date"] in spike_map:
            s["explanation"] = spike_map[s["date"]].get("explanation", "")
            s["operational_tip"] = spike_map[s["date"]].get("operational_tip", "")

    ai_insights = generate_dashboard_ai_insights(
        primary_kpis,
        best_performers,
        charts["best_products_by_revenue"],
        charts["category_revenue_share"],
        top_schools_table,
        best_performers["top_customer"]
    )

    selected_cycle = get_selected_cycle()
    all_cycles = get_all_cycles()

    return render_template(
        'fancy/fancy_dashboard.html',
        # Strict Primary KPIs (as requested by user)
        primary_kpis=primary_kpis,
        total_bookings=primary_kpis["total_bookings"],
        total_revenue=primary_kpis["total_revenue"],
        avg_revenue_per_booking=primary_kpis["avg_revenue_per_booking"],
        unique_customers=primary_kpis["unique_customers"],
        total_units_rented=primary_kpis["total_units_rented"],
        # Best Performers
        best_performers=best_performers,
        # Charts Datasets
        charts=charts,
        best_products_by_revenue=charts["best_products_by_revenue"],
        category_revenue_share=charts["category_revenue_share"],
        weekly_pickup_pattern=charts["weekly_pickup_pattern"],
        monthly_revenue_data=charts["monthly_revenue"],
        daily_velocity=charts["daily_velocity"],
        # Spikes & Anomalies
        spikes=spikes,
        # Tables
        top_schools_table=top_schools_table,
        top_customers_table=top_customers_table,
        # Forecast Calendar with AI Advisory
        forecast_calendar=forecast_calendar,
        # Dashboard-Wide AI Strategic Insights
        ai_insights=ai_insights,
        # Cycle info
        selected_cycle=selected_cycle,
        all_cycles=all_cycles,
        total_stock=total_stock,
        advanced_stats=metrics.get("advanced_stats", {})
    )


# -------------------------------------------------------------
# DRILL-DOWN REST API ENDPOINT (Category / Product / School / Customer / Month / Day / Date / Cycle)
# -------------------------------------------------------------
import urllib.parse

@fancy.route('/api/fancy/drilldown/<entity_type>', defaults={'entity_id': ''}, strict_slashes=False)
@fancy.route('/api/fancy/drilldown/<entity_type>/<path:entity_id>', strict_slashes=False)
def api_fancy_drilldown(entity_type, entity_id):
    if not session.get('logged_in'):
        return jsonify({"status": "error", "error": "Unauthorized"}), 401

    try:
        # Fallback to query param 'id' if entity_id is not in path
        if not entity_id or not entity_id.strip():
            entity_id = request.args.get('id', '').strip()

        if entity_id:
            entity_id = urllib.parse.unquote(entity_id).strip()

        collection = get_selected_collection()
        bookings = list(collection.find())

        entity_type_clean = entity_type.strip().lower()
        entity_data = None

        # Provide sensible default if entity_id is still empty
        if not entity_id:
            if entity_type_clean == 'cycle':
                entity_id = 'current'
            elif entity_type_clean == 'category':
                items, _ = normalize_cycle_bookings(bookings)
                cats = [it["canonical_category"] for it in items if it.get("canonical_category")]
                if cats:
                    entity_id = Counter(cats).most_common(1)[0][0]
                else:
                    return jsonify({"status": "error", "error": "No categories available in the active cycle"}), 404
            elif entity_type_clean == 'month':
                # Default to latest month
                for b in bookings:
                    d = safe_parse_date(b.get("start_date") or b.get("booking_date"))
                    if d:
                        entity_id = f"{d.year}-{d.month:02d}"
                        break
                if not entity_id:
                    entity_id = datetime.now().strftime("%Y-%m")
            else:
                return jsonify({"status": "error", "error": f"{entity_type.capitalize()} ID parameter is required"}), 400

        if entity_type_clean == 'category':
            entity_data = get_category_drilldown(entity_id, bookings)
        elif entity_type_clean == 'product':
            entity_data = get_product_drilldown(entity_id, bookings)
        elif entity_type_clean == 'school':
            entity_data = get_school_drilldown(entity_id, bookings)
        elif entity_type_clean == 'customer':
            all_bookings = []
            for cycle in get_all_cycles():
                c_col = db[cycle["collection_name"]]
                all_bookings.extend(list(c_col.find()))
            entity_data = get_customer_drilldown(entity_id, bookings, all_bookings)
        elif entity_type_clean == 'month':
            entity_data = get_month_drilldown(entity_id, bookings)
        elif entity_type_clean == 'day':
            entity_data = get_day_drilldown(entity_id, bookings)
        elif entity_type_clean == 'date':
            entity_data = get_date_drilldown(entity_id, bookings)
        elif entity_type_clean == 'cycle':
            entity_data = get_cycle_drilldown(bookings)
        else:
            return jsonify({"status": "error", "error": f"Invalid entity type '{entity_type}'"}), 400

        if not entity_data:
            return jsonify({"status": "error", "error": f"{entity_type.capitalize()} '{entity_id}' not found"}), 404

        # Generate AI insights (Completely disabled for customer)
        if entity_type_clean == 'customer':
            ai_insights = []
        else:
            ai_insights = generate_drilldown_ai_insights(entity_type.capitalize(), entity_data)

        return jsonify({
            "status": "success",
            "data": entity_data,
            "ai_insights": ai_insights
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "error": f"Failed to retrieve {entity_type} drilldown: {str(e)}"}), 500


# -------------------------------------------------------------
# ON-DEMAND AI REFRESH API
# -------------------------------------------------------------
@fancy.route('/api/fancy/refresh_ai', methods=['POST'])
def api_fancy_refresh_ai():
    if not session.get('logged_in'):
        return jsonify({"error": "Unauthorized"}), 401

    # Invalidate AI cache in MongoDB
    db["fancy_ai_cache"].delete_many({})
    return jsonify({"status": "success", "message": "AI cache refreshed successfully"})


from io import BytesIO
from flask import send_file
from openpyxl import Workbook

@fancy.route('/download_dashboard_excel')
def download_dashboard_excel():

    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))

    collection = get_selected_collection()
    current_bookings = list(collection.find())

    items, summary = normalize_cycle_bookings(current_bookings)

    wb = Workbook()

    # Sheet 1: Normalized Line Items
    ws1 = wb.active
    ws1.title = "Normalized Rentals"
    ws1.append([
        "Booking ID",
        "Date",
        "Customer Name",
        "School",
        "Canonical Product",
        "Category",
        "Units Rented",
        "Allocated Revenue (Rs)"
    ])
    for it in items:
        ws1.append([
            it["booking_id"],
            it["booking_date"],
            it["customer_name"],
            it["school_clean"],
            it["canonical_product"],
            it["canonical_category"],
            it["units"],
            it["allocated_revenue"]
        ])

    # Sheet 2: Category Summary
    ws2 = wb.create_sheet("Category Summary")
    ws2.append(["Category", "Units Rented", "Total Revenue (Rs)"])
    cat_rev = Counter()
    cat_units = Counter()
    for it in items:
        cat_rev[it["canonical_category"]] += it["allocated_revenue"]
        cat_units[it["canonical_category"]] += it["units"]
    for c, rev in cat_rev.most_common():
        ws2.append([c, cat_units[c], round(rev, 2)])

    # Sheet 3: Product Summary
    ws3 = wb.create_sheet("Product Summary")
    ws3.append(["Product Name", "Category", "Units Rented", "Total Revenue (Rs)"])
    prod_rev = Counter()
    prod_units = Counter()
    prod_cat = {}
    for it in items:
        p = it["canonical_product"]
        prod_rev[p] += it["allocated_revenue"]
        prod_units[p] += it["units"]
        prod_cat[p] = it["canonical_category"]
    for p, rev in prod_rev.most_common():
        ws3.append([p, prod_cat.get(p, "General"), prod_units[p], round(rev, 2)])

    # Sheet 4: School Leaderboard (Current Cycle)
    ws4 = wb.create_sheet("School Leaderboard")
    ws4.append(["Rank", "School Name", "Bookings", "Total Revenue (Rs)"])
    school_rev = Counter()
    school_cnt = Counter()
    for it in items:
        if not it["is_walkin"]:
            school_rev[it["school_clean"]] += it["allocated_revenue"]
            school_cnt[it["school_clean"]] += 1
    for rank, (s, rev) in enumerate(school_rev.most_common(), 1):
        ws4.append([rank, s, school_cnt[s], round(rev, 2)])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="fancy_dashboard_report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@fancy.route("/fancy_inventory", methods=["GET", "POST"])
def fancy_inventory():
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))
    if request.method == "POST":
        name = request.form.get("name")
        color = request.form.get("color")
        category = request.form.get("category")

        size_names = request.form.getlist("size_name[]")
        size_qtys = request.form.getlist("size_qty[]")

        sizes = {}
        for s, q in zip(size_names, size_qtys):
            if s.strip() and q.strip():
                sizes[s.strip()] = int(q)

        finventory.insert_one({
            "name": name,
            "color": color,
            "category": category,
            "sizes": sizes
        })

        return redirect(url_for("fancy.fancy_inventory"))

    products = list(finventory.find())
    return render_template("fancy/fancy_inventory.html", products=products)


@fancy.route("/fancy_inventory/update/<id>", methods=["POST"])
def update_fancy_inventory(id):
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))
    name = request.form.get("name")
    color = request.form.get("color")
    category = request.form.get("category")

    size_names = request.form.getlist("size_name[]")
    size_qtys = request.form.getlist("size_qty[]")

    sizes = {}
    for s, q in zip(size_names, size_qtys):
        if s.strip() and q.strip():
            sizes[s.strip()] = int(q)

    finventory.update_one(
        {"_id": ObjectId(id)},
        {"$set": {
            "name": name,
            "color": color,
            "category": category,
            "sizes": sizes
        }}
    )

    return redirect(url_for("fancy.fancy_inventory"))


@fancy.route("/fancy_inventory/delete/<id>", methods=["POST"])
def delete_fancy_inventory(id):
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))
    finventory.delete_one({"_id": ObjectId(id)})
    return redirect(url_for("fancy.fancy_inventory"))

@fancy.route('/fancy_profile')
def fancy_profile():

    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))

    mobile = request.args.get('mobile')

    if not mobile:
        return "Mobile number missing", 400

    # Customer master
    customer = fcustomers.find_one({
        "mobile": mobile
    })

    if not customer:
        return "Customer not found", 404

    all_bookings = []

    # ---------- FETCH FROM ALL CYCLES ----------
    for cycle in get_all_cycles():

        collection = db[
            cycle["collection_name"]
        ]

        bookings = list(
            collection.find({
                "mobile": mobile
            })
        )

        for b in bookings:

            b["season"] = cycle["name"]
            b["cycle_id"] = str(
                cycle["_id"]
            )

            # Normalize start_date
            sd = b.get("start_date")

            if isinstance(sd, datetime):

                b["start_date"] = sd.strftime(
                    "%d-%m-%Y"
                )

            elif isinstance(sd, str):

                formats = [
                    "%Y-%m-%d",
                    "%d-%m-%Y",
                    "%d-%m-%y"
                ]

                for fmt in formats:
                    try:
                        b["start_date"] = (
                            datetime.strptime(
                                sd,
                                fmt
                            ).strftime(
                                "%d-%m-%Y"
                            )
                        )
                        break
                    except:
                        pass

            # Normalize end_date
            ed = b.get("end_date")

            if isinstance(ed, datetime):

                b["end_date"] = ed.strftime(
                    "%d-%m-%Y"
                )

            elif isinstance(ed, str):

                formats = [
                    "%Y-%m-%d",
                    "%d-%m-%Y",
                    "%d-%m-%y"
                ]

                for fmt in formats:
                    try:
                        b["end_date"] = (
                            datetime.strptime(
                                ed,
                                fmt
                            ).strftime(
                                "%d-%m-%Y"
                            )
                        )
                        break
                    except:
                        pass

        all_bookings.extend(bookings)

    # ---------- SORT LATEST FIRST ----------
    all_bookings.sort(
        key=lambda x: x.get(
            "timestamp",
            datetime.min
        ),
        reverse=True
    )

    total_spent = sum(
        b.get("price", 0)
        for b in all_bookings
    )

    return render_template(
        "fancy/fancy_profile.html",
        customer=customer,
        bookings=all_bookings,
        total_spent=total_spent
    )

@fancy.route("/fancy_cycles")
def fancy_cycles_page():

    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))

    cycles = get_all_cycles()

    return render_template(
        "fancy/fancy_cycles.html",
        cycles=cycles
    )

@fancy.route("/fancy_cycles/create", methods=["POST"])
def create_fancy_cycle_route():

    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))

    name = request.form.get("name")
    collection_name = request.form.get("collection_name")

    create_cycle(
        name,
        collection_name
    )

    return redirect(
        url_for("fancy.fancy_cycles_page")
    )


@fancy.route("/fancy_cycles/select/<cycle_id>", methods=["GET", "POST"])
def select_fancy_cycle_id(cycle_id):
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))

    set_selected_cycle(cycle_id)
    return redirect("/fancy_admin")


@fancy.route("/fancy_cycles/select", methods=["GET", "POST"])
def select_fancy_cycle():
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))

    cycle_id = request.form.get("cycle_id") or request.args.get("cycle_id")
    if cycle_id:
        set_selected_cycle(cycle_id)

    return redirect("/fancy_admin")

@fancy.route("/fancy_cycles/end", methods=["POST"])
@fancy.route("/fancy_cycles/end/<cycle_id>", methods=["GET", "POST"])
def end_fancy_cycle_route(cycle_id=None):
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))

    if request.method == "POST":
        target_id = request.form.get("cycle_id") or cycle_id
        password = request.form.get("password")

        if password != ADMIN_PASS:
            flash("❌ Authentication failed: Invalid Admin Password!", "error")
            return redirect(url_for("fancy.fancy_cycles_page"))

        if end_cycle(target_id):
            flash("✅ Active Fancy cycle ended successfully.", "success")
        else:
            flash("❌ Could not end cycle.", "error")
        return redirect(url_for("fancy.fancy_cycles_page"))

    flash("⚠️ Password confirmation required to end a cycle.", "error")
    return redirect(url_for("fancy.fancy_cycles_page"))


@fancy.route("/fancy_cycles/reactivate", methods=["POST"])
@fancy.route("/fancy_cycles/reactivate/<cycle_id>", methods=["GET", "POST"])
def reactivate_fancy_cycle_route(cycle_id=None):
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))

    if request.method == "POST":
        target_id = request.form.get("cycle_id") or cycle_id
        password = request.form.get("password")

        if password != ADMIN_PASS:
            flash("❌ Authentication failed: Invalid Admin Password!", "error")
            return redirect(url_for("fancy.fancy_cycles_page"))

        success, msg = reactivate_cycle(target_id)
        if success:
            flash(msg, "success")
        else:
            flash(f"❌ {msg}", "error")
        return redirect(url_for("fancy.fancy_cycles_page"))

    flash("⚠️ Password confirmation required to reactivate a cycle.", "error")
    return redirect(url_for("fancy.fancy_cycles_page"))


@fancy.route("/fancy_cycles/unlock/<cycle_id>", methods=["POST"])
def unlock_cycle(cycle_id):

    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))

    entered_id = request.form.get('id')
    entered_pass = request.form.get('password')

    if entered_id != ADMIN_ID or entered_pass != ADMIN_PASS:
        flash("❌ Invalid credentials!", "error")
        return redirect("/fancy_admin")

    fancy_cycles.update_one(
        {"_id": ObjectId(cycle_id)},
        {
            "$set": {
                "edit_override": True
            }
        }
    )

    flash("🔓 Cycle unlocked successfully!", "success")
    return redirect("/fancy_admin")


@fancy.route(
    "/fancy_cycles/lock/<cycle_id>"
)
def lock_cycle(cycle_id):

    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))

    fancy_cycles.update_one(
        {"_id": ObjectId(cycle_id)},
        {
            "$set": {
                "edit_override": False
            }
        }
    )

    flash("🔒 Cycle locked successfully!", "success")
    return redirect("/fancy_admin")

KNOWN_LOCALITIES_FANCY = [
    "Lambha", "Dehgam", "Patan", "Palanpur", "Unjha", "Visnagar", "Mehsana", "Kalol", "Chhatral", "Kadi", "Himmatnagar", "Gandhinagar",
    "Nadiad", "Anand", "Bakrol", "Vadtal", "Vadodara", "Surat", "Rajkot", "Kheda", "Sanand", "Dholka", "Bavla",
    "Aslali", "Bareja", "Changodar", "Moraiya",
    "Nava Vadaj", "Vadaj", "Jay Hind", "Arbuda Nagar", "Haridarshan", "Hathijan", "Vivekanand Nagar",
    "Saijpur Bogha", "Saijpur", "Jivraj Park", "South Bopal", "Bopal", "Ghatlodiya", "Gordhanwadi", "Maniyasa", "Laxminarayan",
    "Jashodanagar", "Jashoda Nagar", "Bhadwat Nagar", "Prahlad Nagar", "Prahladnagar", "Prerna Tirth", "Satelite",
    "New Maninagar", "New Vatva", "Shahwadi", "Motipura", "Sureliya", "Khodiyar Nagar", "Rajendra Park", "Saheed Circle", "Aman Nagar",
    "Vastral", "Maninagar", "Khokhra", "Isanpur", "Amraiwadi", "Ghodasar",
    "Vatva", "Odhav", "Hatkeshwar", "CTM", "Nikol", "Ramol", "Narol",
    "Bapunagar", "Saraspur", "Asarwa", "Shahibaug", "Naroda", "Rakhial",
    "Sarangpur", "Kalupur", "Astodia", "Raipur", "Lal Darwaja", "Geeta Mandir",
    "Shahpur", "Dani Limda", "Navrangpura", "Satellite", "Vastrapur", "Bodakdev",
    "Thaltej", "Sola", "Gota", "Ghatlodia", "Naranpura", "Paldi", "Vasna",
    "Ranip", "Sabarmati", "Chandkheda"
]

AREA_COORDINATES_FANCY = {
    "Lambha": [22.9238, 72.5843],
    "Dehgam": [23.1670, 72.8120],
    "Patan": [23.8493, 72.1266],
    "Palanpur": [24.1724, 72.4346],
    "Unjha": [23.8043, 72.3942],
    "Visnagar": [23.6961, 72.5484],
    "Mehsana": [23.5880, 72.3693],
    "Kalol": [23.2393, 72.4962],
    "Chhatral": [23.2800, 72.4500],
    "Kadi": [23.3000, 72.3300],
    "Himmatnagar": [23.5979, 72.9698],
    "Gandhinagar": [23.2156, 72.6369],
    "Nadiad": [22.6916, 72.8634],
    "Anand": [22.5645, 72.9289],
    "Bakrol": [22.5480, 72.9350],
    "Vadtal": [22.5920, 72.8880],
    "Vadodara": [22.3072, 73.1812],
    "Surat": [21.1702, 72.8311],
    "Rajkot": [22.3039, 70.8022],
    "Kheda": [22.7500, 72.6800],
    "Sanand": [22.9910, 72.3810],
    "Dholka": [22.7200, 72.4700],
    "Bavla": [22.8300, 72.3600],
    "Aslali": [22.9210, 72.6010],
    "Bareja": [22.8850, 72.6050],
    "Changodar": [22.9230, 72.4410],
    "Moraiya": [22.9150, 72.4350],
    "Nava Vadaj": [23.0640, 72.5690],
    "Vadaj": [23.0640, 72.5690],
    "Jay Hind": [22.9920, 72.5980],
    "Arbuda Nagar": [23.0250, 72.6630],
    "Haridarshan": [23.0450, 72.6680],
    "Hathijan": [22.9280, 72.6390],
    "Vivekanand Nagar": [22.9280, 72.6390],
    "Saijpur Bogha": [23.0640, 72.6280],
    "Saijpur": [23.0640, 72.6280],
    "Jivraj Park": [23.0010, 72.5410],
    "South Bopal": [23.0300, 72.4640],
    "Bopal": [23.0300, 72.4640],
    "Ghatlodiya": [23.0682, 72.5358],
    "Gordhanwadi": [22.9980, 72.5920],
    "Maniyasa": [22.9976, 72.6009],
    "Laxminarayan": [22.9554, 72.6240],
    "Jashodanagar": [22.9850, 72.6250],
    "Jashoda Nagar": [22.9850, 72.6250],
    "Bhadwat Nagar": [22.9910, 72.6080],
    "Prahlad Nagar": [23.0125, 72.5118],
    "Prahladnagar": [23.0125, 72.5118],
    "Prerna Tirth": [23.0300, 72.5176],
    "Satelite": [23.0300, 72.5176],
    "Vastral": [23.0041, 72.6617],
    "Maninagar": [22.9976, 72.6009],
    "New Maninagar": [22.9850, 72.6150],
    "Khokhra": [22.9983, 72.6167],
    "Isanpur": [22.9731, 72.5976],
    "Amraiwadi": [23.0039, 72.6288],
    "Ghodasar": [22.9815, 72.6094],
    "Vatva": [22.9554, 72.6240],
    "New Vatva": [22.9480, 72.6310],
    "Odhav": [23.0232, 72.6698],
    "Hatkeshwar": [23.0012, 72.6225],
    "CTM": [22.9908, 72.6321],
    "Nikol": [23.0483, 72.6717],
    "Ramol": [22.9840, 72.6582],
    "Narol": [22.9634, 72.5891],
    "Shahwadi": [22.9570, 72.5780],
    "Motipura": [22.9610, 72.5820],
    "Sureliya": [23.0010, 72.6510],
    "Khodiyar Nagar": [23.0390, 72.6350],
    "Rajendra Park": [23.0210, 72.6610],
    "Aman Nagar": [23.0240, 72.6650],
    "Saheed Circle": [23.0490, 72.6730],
    "Bapunagar": [23.0371, 72.6231],
    "Saraspur": [23.0298, 72.6080],
    "Asarwa": [23.0494, 72.6033],
    "Shahibaug": [23.0560, 72.5925],
    "Naroda": [23.0725, 72.6656],
    "Rakhial": [23.0180, 72.6210],
    "Sarangpur": [23.0215, 72.5990],
    "Kalupur": [23.0260, 72.5950],
    "Astodia": [23.0170, 72.5910],
    "Raipur": [23.0185, 72.5940],
    "Lal Darwaja": [23.0240, 72.5810],
    "Geeta Mandir": [23.0110, 72.5880],
    "Shahpur": [23.0350, 72.5780],
    "Dani Limda": [22.9950, 72.5810],
    "Navrangpura": [23.0366, 72.5611],
    "Satellite": [23.0300, 72.5176],
    "Vastrapur": [23.0350, 72.5293],
    "Bodakdev": [23.0410, 72.5115],
    "Thaltej": [23.0500, 72.5070],
    "Sola": [23.0680, 72.5180],
    "Gota": [23.0970, 72.5310],
    "Ghatlodia": [23.0682, 72.5358],
    "Naranpura": [23.0520, 72.5530],
    "Paldi": [23.0120, 72.5620],
    "Vasna": [22.9980, 72.5520],
    "Ranip": [23.0800, 72.5710],
    "Sabarmati": [23.0845, 72.5802],
    "Chandkheda": [23.1114, 72.5835]
}


@fancy.route("/fancy-customers")
def fancy_customers():

    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))

    search = request.args.get("search", "").strip()

    query = {}

    if search:
        query = {
            "$or": [
                {"name": {"$regex": search, "$options": "i"}},
                {"mobile": {"$regex": search, "$options": "i"}},
                {"school": {"$regex": search, "$options": "i"}},
                {"address": {"$regex": search, "$options": "i"}}
            ]
        }

    all_customers = list(fcustomers.find().sort("updated_at", -1))

    # Dynamic Custom Localities Merge
    from website.general.db import custom_localities
    active_localities = list(KNOWN_LOCALITIES_FANCY)
    active_coords = dict(AREA_COORDINATES_FANCY)
    try:
        for cloc in custom_localities.find():
            cname = cloc.get("name")
            clat = cloc.get("lat")
            clng = cloc.get("lng")
            if cname and clat is not None and clng is not None:
                if cname not in active_localities:
                    active_localities.insert(0, cname)
                active_coords[cname] = [float(clat), float(clng)]
    except Exception:
        pass

    # Area Strength Analytics & Verified Locality Mapping
    from website.general.utils import resolve_customer_locality

    area_counts = {}
    total_with_addr = 0
    unmapped_count = 0

    for c in all_customers:
        matched_loc = resolve_customer_locality(c, active_localities)
        if matched_loc:
            c["mapped_locality"] = matched_loc
            area_counts[matched_loc] = area_counts.get(matched_loc, 0) + 1
            total_with_addr += 1
        else:
            c["mapped_locality"] = ""
            unmapped_count += 1

    sorted_areas = sorted(area_counts.items(), key=lambda x: x[1], reverse=True)
    top_areas = []
    map_localities = []

    for loc, count in sorted_areas:
        pct = round((count / max(total_with_addr, 1)) * 100, 1)
        item = {"area": loc, "count": count, "percentage": pct}
        if len(top_areas) < 5:
            top_areas.append(item)
        if loc in active_coords:
            item_map = dict(item)
            item_map["lat"] = active_coords[loc][0]
            item_map["lng"] = active_coords[loc][1]
            map_localities.append(item_map)

    id_to_loc = {str(c["_id"]): c.get("mapped_locality", "") for c in all_customers}

    if search:
        customers = list(fcustomers.find(query).sort("updated_at", -1))
        for c in customers:
            c["mapped_locality"] = id_to_loc.get(str(c["_id"]), "")
    else:
        customers = all_customers

    return render_template(
        "fancy/fancy_customers.html",
        customers=customers,
        total_count=len(all_customers),
        total_with_addr=total_with_addr,
        unmapped_count=unmapped_count,
        top_areas=top_areas,
        map_localities=map_localities,
        area_map_data=map_localities,
        search=search
    )

@fancy.route(
    "/fancy-customer/<customer_id>",
    methods=["GET", "POST"]
)
def fancy_customer(customer_id):

    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))

    customer = fcustomers.find_one(
        {"_id": ObjectId(customer_id)}
    )

    if not customer:
        return "Customer Not Found"

    if request.method == "POST":

        fcustomers.update_one(
            {"_id": ObjectId(customer_id)},
            {
                "$set": {
                    "name": request.form.get("name"),
                    "mobile": request.form.get("mobile"),
                    "school": request.form.get("school"),
                    "address": request.form.get("address"),
                    "updated_at": datetime.utcnow()
                }
            }
        )

        flash("Customer Updated", "success")

        return redirect(
            url_for("fancy.fancy_customers")
        )

    return render_template(
        "fancy/fancy_customer_edit.html",
        customer=customer
    )

@fancy.route(
    "/fancy-customer/delete/<customer_id>",
    methods=["POST"]
)
def delete_fancy_customer(customer_id):

    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))

    fcustomers.delete_one(
        {"_id": ObjectId(customer_id)}
    )

    flash("Customer Deleted", "success")

    return redirect(
        url_for("fancy.fancy_customers")
    )


# ------------------ FANCY ACTION LOGS ------------------
def log_fancy_action(name, mobile, action, details):
    from website.general.utils import get_ist_now
    selected_cycle = get_selected_cycle()
    if not selected_cycle:
        return

    collection_name = selected_cycle.get("collection_name")
    if not collection_name:
        return

    logs_col = db[f"{collection_name}_logs"]
    now = get_ist_now()

    log_entry = {
        "name": name or "",
        "mobile": mobile or "",
        "action": action,
        "details": details,
        "date_stamp": now.strftime("%d/%m/%Y"),
        "time_stamp": now.strftime("%H:%M:%S"),
        "timestamp": now
    }

    try:
        logs_col.insert_one(log_entry)
    except Exception:
        pass


@fancy.route("/fancy_logs")
def fancy_logs():
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))

    from website.general.utils import format_log_timestamp
    selected_cycle = get_selected_cycle()
    logs = []
    if selected_cycle:
        collection_name = selected_cycle.get("collection_name")
        if collection_name:
            logs_col = db[f"{collection_name}_logs"]
            raw_logs = list(logs_col.find().sort("timestamp", -1))
            for log in raw_logs:
                d_str, t_str, sort_ts = format_log_timestamp(
                    log.get("timestamp"),
                    log.get("date_stamp", ""),
                    log.get("time_stamp", "")
                )
                log["date_stamp"] = d_str
                log["time_stamp"] = t_str
                log["sort_ts"] = sort_ts
                logs.append(log)

    return render_template(
        "fancy/fancy_logs.html",
        logs=logs,
        selected_cycle=selected_cycle
    )


@fancy.route("/fancy_logs/api")
def fancy_logs_api():
    if not session.get('logged_in'):
        return jsonify({"success": False, "message": "Unauthorized"}), 401

    from website.general.utils import format_log_timestamp
    selected_cycle = get_selected_cycle()
    logs_data = []
    if selected_cycle:
        collection_name = selected_cycle.get("collection_name")
        if collection_name:
            logs_col = db[f"{collection_name}_logs"]
            raw_logs = list(logs_col.find().sort("timestamp", -1))
            for log in raw_logs:
                d_str, t_str, sort_ts = format_log_timestamp(
                    log.get("timestamp"),
                    log.get("date_stamp", ""),
                    log.get("time_stamp", "")
                )
                logs_data.append({
                    "id": str(log.get("_id", "")),
                    "name": log.get("name", "") or "—",
                    "mobile": log.get("mobile", "") or "—",
                    "action": log.get("action", ""),
                    "details": log.get("details", ""),
                    "date_stamp": d_str,
                    "time_stamp": t_str,
                    "sort_ts": sort_ts
                })

    return jsonify({"success": True, "logs": logs_data, "cycle_name": selected_cycle.get("name") if selected_cycle else ""})



@fancy.route("/fancy_logs/clear", methods=["POST"])
def clear_fancy_logs():
    if not session.get('logged_in'):
        return jsonify({"success": False, "message": "Unauthorized"}), 401

    data = request.json or request.form
    password = data.get("password", "").strip()

    if password != ADMIN_PASS:
        return jsonify({"success": False, "message": "❌ Authentication failed: Invalid Admin Password!"}), 400

    selected_cycle = get_selected_cycle()
    if not selected_cycle:
        return jsonify({"success": False, "message": "No Fancy cycle selected."}), 400

    collection_name = selected_cycle.get("collection_name")
    if not collection_name:
        return jsonify({"success": False, "message": "Invalid cycle collection."}), 400

    logs_col = db[f"{collection_name}_logs"]
    logs_col.delete_many({})

    try:
        log_fancy_action("Admin", "", "clear_logs", f"Cleared all action logs for Fancy cycle '{selected_cycle.get('name')}'.")
    except Exception:
        pass

    return jsonify({"success": True, "message": "✅ All Fancy action logs cleared successfully!"})